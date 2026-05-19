import pyodbc
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

# --- Logging ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
    handlers=[
        logging.FileHandler('customer_trend_report.log', mode='w', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# --- Configuration ---
# MSSQL connection details
server = '192.168.68.54'
database = 'pcay'
username = 'paradise'
password = 'paradise1'

# Years configuration
YEARS_FULL_YEAR = [2023, 2024, 2025]  # Years to show both full year and YTD
YEAR_YTD_ONLY = [2026]  # Years to show YTD only
all_years = YEARS_FULL_YEAR + YEAR_YTD_ONLY

# --- Dynamic Year Calculation ---
current_date = datetime.now()
current_year = current_date.year
day_of_year = current_date.timetuple().tm_yday
years_str = ", ".join(map(str, all_years))

# --- Dynamic SQL Query Generation ---
# Build the parts of the query that depend on the years
purchase_sum_cols = []

# Add full year columns for 2023, 2024, 2025
for year in YEARS_FULL_YEAR:
    purchase_sum_cols.append(f"SUM(CASE WHEN YEAR(COALESCE(TxnDate, ShipDate)) = {year} THEN InvoiceLineAmount ELSE 0 END) AS Purchased{year}FullYear")

# Add YTD columns for all years
for year in all_years:
    purchase_sum_cols.append(f"SUM(CASE WHEN YEAR(COALESCE(TxnDate, ShipDate)) = {year} AND DATEPART(dy, COALESCE(TxnDate, ShipDate)) <= {day_of_year} THEN InvoiceLineAmount ELSE 0 END) AS Purchased{year}YTD")

purchase_sum_statement = ",\n        ".join(purchase_sum_cols)

final_select_cols = []

# Add full year columns to final select
for year in YEARS_FULL_YEAR:
    final_select_cols.append(f"cs.Purchased{year}FullYear")

# Add YTD columns to final select
for year in all_years:
    final_select_cols.append(f"cs.Purchased{year}YTD")

final_select_statement = ",\n    ".join(final_select_cols)

query = f'''
WITH CustomerSales AS (
    SELECT
        CustomerRefListID AS CustomerID,
        {purchase_sum_statement},
        SUM(InvoiceLineAmount) AS TotalSales
    FROM dbo.InvoiceLine
    WHERE
        CustomerRefListID IS NOT NULL
        AND YEAR(COALESCE(TxnDate, ShipDate)) IN ({years_str})
    GROUP BY CustomerRefListID
)
SELECT
    c.Name AS CustomerName,
    {final_select_statement},
    cs.TotalSales,
    CONCAT_WS(' ', c.FirstName, c.LastName) AS ContactName,
    c.Email,
    c.Phone AS PhoneNumber,
    CONCAT_WS(', ', c.BillAddressAddr1, c.BillAddressAddr2, c.BillAddressCity, c.BillAddressState, c.BillAddressPostalCode) AS BillingAddress
FROM CustomerSales cs
JOIN dbo.Customer c ON cs.CustomerID = c.ListID
WHERE cs.TotalSales > 0
ORDER BY c.Name;
'''

# --- Database Connection and Query Execution ---
try:
    logger.info('Starting customer trend report run')
    conn_str = (
        f'DRIVER={{ODBC Driver 18 for SQL Server}};'
        f'SERVER={server};'
        f'DATABASE={database};'
        f'UID={username};'
        f'PWD={password};'
        f'TrustServerCertificate=yes;'
    )
    logger.info('Connecting to SQL Server')
    conn = pyodbc.connect(conn_str)
    
    # Run query and fetch results
    logger.info('Running main query')
    results = pd.read_sql(query, conn)
    logger.info('Query returned %d rows', len(results))
    
    # --- Report Generation ---
    print(f'\nCustomer Purchase Report for 2023-2026')
    print('-' * 120)
    
    # Format columns as "Full Year/YTD" for display
    for year in YEARS_FULL_YEAR:
        col_full = f'Purchased{year}FullYear'
        col_ytd = f'Purchased{year}YTD'
        # Create formatted column
        results[f'{year}'] = results.apply(lambda row: f"{row[col_full]:,.0f} / {row[col_ytd]:,.0f}", axis=1)
        # Drop the original columns
        results.drop(columns=[col_full, col_ytd], inplace=True)
    
    # Format YTD only columns for 2026
    for year in YEAR_YTD_ONLY:
        col_ytd = f'Purchased{year}YTD'
        results[f'{year}'] = results[col_ytd].apply(lambda x: f"{x:,.0f}")
        results.drop(columns=[col_ytd], inplace=True)
    
    # Drop TotalSales as we're showing individual years now
    results.drop(columns=['TotalSales'], inplace=True)
    
    print(results.to_string(index=False))
    print('-' * 120)

    # Calculate and print summary statistics
    print(f'\nSummary:')
    for year in YEARS_FULL_YEAR:
        col_full = f'Purchased{year}FullYear'
        col_ytd = f'Purchased{year}YTD'
        total_full = pd.read_sql(
            f"SELECT SUM(CASE WHEN YEAR(COALESCE(TxnDate, ShipDate)) = {year} THEN InvoiceLineAmount ELSE 0 END) as total FROM dbo.InvoiceLine WHERE CustomerRefListID IS NOT NULL",
            conn
        ).iloc[0, 0]
        total_ytd = pd.read_sql(
            f"SELECT SUM(CASE WHEN YEAR(COALESCE(TxnDate, ShipDate)) = {year} AND DATEPART(dy, COALESCE(TxnDate, ShipDate)) <= {day_of_year} THEN InvoiceLineAmount ELSE 0 END) as total FROM dbo.InvoiceLine WHERE CustomerRefListID IS NOT NULL",
            conn
        ).iloc[0, 0]
        print(f'{year}: ${total_full:,.2f} / ${total_ytd:,.2f}')
    
    for year in YEAR_YTD_ONLY:
        total_ytd = pd.read_sql(
            f"SELECT SUM(CASE WHEN YEAR(COALESCE(TxnDate, ShipDate)) = {year} AND DATEPART(dy, COALESCE(TxnDate, ShipDate)) <= {day_of_year} THEN InvoiceLineAmount ELSE 0 END) as total FROM dbo.InvoiceLine WHERE CustomerRefListID IS NOT NULL",
            conn
        ).iloc[0, 0]
        print(f'{year}: ${total_ytd:,.2f}')

    # --- Excel Export and Formatting ---
    excel_path = 'customer_change_report_last_6_years.xlsx'
    logger.info('Writing Excel output to %s', excel_path)
    
    # Reorder columns: CustomerName, Year columns, empty space, ContactName, Email, PhoneNumber, BillingAddress
    year_cols = [str(year) for year in YEARS_FULL_YEAR + YEAR_YTD_ONLY]
    
    # Insert empty columns for spacing
    results.insert(1, '  ', '')  # Empty column after CustomerName
    contact_name_index = len(['CustomerName', '  '] + year_cols)
    results.insert(contact_name_index, ' ', '')  # Empty column before ContactName
    
    ordered_cols = ['CustomerName', '  '] + year_cols + [' ', 'ContactName', 'Email', 'PhoneNumber', 'BillingAddress']
    results = results[ordered_cols]

    results.to_excel(excel_path, index=False)

    # Load workbook for formatting
    logger.info('Loading workbook for formatting')
    wb = load_workbook(excel_path)
    ws = wb.active

    # Define fills
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    orange_fill = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')

    # Find the header names to locate columns dynamically
    headers = [cell.value for cell in ws[1]]
    
    # Find the columns for the years
    year_columns_indices = [i for i, h in enumerate(headers) if h and str(h) in year_cols]

    for row in range(2, ws.max_row + 1):
        # Determine which cells to highlight
        highlight_row = False
        purchases = {}
        for col_idx in year_columns_indices:
            cell_value = ws.cell(row=row, column=col_idx + 1).value
            # Parse the value to check if it's significant (value > 2000)
            if isinstance(cell_value, str):
                try:
                    # Extract first number from "Full/YTD" format
                    first_val = float(cell_value.split('/')[0].replace('$', '').replace(',', ''))
                    purchases[col_idx + 1] = first_val
                    if first_val > 2000:
                        highlight_row = True
                except:
                    pass
            elif isinstance(cell_value, (int, float)):
                purchases[col_idx + 1] = cell_value
                if cell_value > 2000:
                    highlight_row = True

        # Apply highlights in the correct order
        if highlight_row:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = orange_fill
        
        if purchases:
            max_year_col_idx = max(purchases, key=purchases.get)
            ws.cell(row=row, column=max_year_col_idx).fill = yellow_fill

    # --- Restore Borders ---
    # Define a thin border style to mimic default gridlines
    thin_border = Border(
        left=Side(style='thin', color='D4D4D4'),
        right=Side(style='thin', color='D4D4D4'),
        top=Side(style='thin', color='D4D4D4'),
        bottom=Side(style='thin', color='D4D4D4')
    )
    # Apply the border to all cells in the used range
    for row in range(1, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border

    # --- Freeze the header row ---
    ws.freeze_panes = 'A2'

    # --- Widen columns A through H ---
    ws.column_dimensions['A'].width = 25  # CustomerName column
    ws.column_dimensions['B'].width = 4   # Empty column (half as wide)
    
    for col_letter in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 15
    
    for col_letter in ['G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col_letter].width = 32

    # Truncate text in CustomerName column by setting horizontal alignment to fill
    for row in range(1, ws.max_row + 1):
        ws['A' + str(row)].alignment = Alignment(horizontal='fill', wrap_text=False)
    


   

    wb.save(excel_path)
    print(f'\nReport saved to {excel_path}.')
    logger.info('Report saved successfully')

except Exception as e:
    logger.exception('An error occurred')
    print(f"An error occurred: {e}")

finally:
    if 'conn' in locals() and conn:
        conn.close()
